"""One-shot inspection of all downloaded raw archives — what's inside each.
Not part of the production pipeline, just helps us decide how to parse.
"""
from pathlib import Path
import zipfile
import openpyxl
import shapefile  # pyshp

RAW = Path(__file__).resolve().parent.parent / "data" / "raw"


def show_zip(zip_path: Path, max_lines: int = 25) -> None:
    print(f"\n=== {zip_path.relative_to(RAW.parent.parent)} ===")
    with zipfile.ZipFile(zip_path) as z:
        names = z.namelist()
        print(f"  total entries: {len(names)}")
        for n in names[:max_lines]:
            info = z.getinfo(n)
            print(f"    {info.file_size:>10}  {n}")
        if len(names) > max_lines:
            print(f"    ... +{len(names) - max_lines} more")


def show_xlsx(xlsx_path: Path) -> None:
    print(f"\n=== {xlsx_path.relative_to(RAW.parent.parent)} ===")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  sheet: {sn}  ({ws.max_row} rows x {ws.max_column} cols)")
        rows_iter = ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True)
        for i, row in enumerate(rows_iter, start=1):
            cells = [str(c)[:30] if c is not None else "" for c in row[:10]]
            print(f"    row {i}: {cells}")
    wb.close()


def peek_shp_in_zip(zip_path: Path) -> None:
    """Pyshp can read .shp inside a .zip directly."""
    print(f"\n=== SHP contents of {zip_path.name} ===")
    with zipfile.ZipFile(zip_path) as z:
        shp_name = next((n for n in z.namelist() if n.lower().endswith(".shp")), None)
        if shp_name is None:
            print("  no .shp inside!")
            return
        print(f"  shapefile: {shp_name}")
        dbf_name = shp_name[:-4] + ".dbf"
        shx_name = shp_name[:-4] + ".shx"
        with z.open(shp_name) as shp, z.open(dbf_name) as dbf, z.open(shx_name) as shx:
            r = shapefile.Reader(shp=shp, dbf=dbf, shx=shx)
            print(f"  records: {len(r)}")
            field_names = [f[0] for f in r.fields if f[0] != "DeletionFlag"]
            print(f"  fields: {field_names}")
            for i, rec in enumerate(r.iterShapeRecords()):
                if i >= 3:
                    break
                attrs = dict(zip(field_names, list(rec.record)))
                print(f"    record {i}: bbox={rec.shape.bbox}  attrs={attrs}")


if __name__ == "__main__":
    show_zip(RAW / "abs_boundaries" / "SAL_2021_AUST_GDA2020_SHP.zip")
    peek_shp_in_zip(RAW / "abs_boundaries" / "SAL_2021_AUST_GDA2020_SHP.zip")
    show_xlsx(RAW / "abs_allocation" / "SAL_2021_AUST.xlsx")
    show_zip(RAW / "abs_census" / "2021_GCP_SAL_for_NSW_short-header.zip", max_lines=15)
    show_xlsx(RAW / "abs_regional_pop" / "32180DS0001_2024-25.xlsx")
    show_zip(RAW / "vg_sales" / "archive.zip", max_lines=15)
    show_zip(RAW / "bocsar" / "SuburbData.zip", max_lines=25)
